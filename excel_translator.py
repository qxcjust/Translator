import logging
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Border, Alignment
from openpyxl.drawing.image import Image
from gl_config import LOG_LEVEL


logging.basicConfig(level=LOG_LEVEL)

def translate_excel(translation_core, file_path, output_path, source_lang, target_lang, task):
    """翻译Excel文件并保持原始格式和结构"""
    logging.info(f"Starting translation of Excel file: {file_path}")
    
    # 加载原始工作簿和创建翻译后的工作簿
    excel_wb = load_workbook(file_path, data_only=True)

    total_sheets = len(excel_wb.sheetnames)

    # TODO openpyxl 无法读取形状 保存后会丢失
    
    for sheet_idx, sheet_name in enumerate(excel_wb.sheetnames, 1):
        # 获取原始Sheet并创建对应的翻译Sheet
        original_sheet = excel_wb[sheet_name]       
        
        # 计算当前Sheet的总单元格数
        total_cells = original_sheet.max_row * original_sheet.max_column
        
        # 翻译单元格内容并复制格式
        for row_idx, row in enumerate(original_sheet.iter_rows(), 1):
            for cell_idx, cell in enumerate(row, 1):
                if cell.value is not None and isinstance(cell.value, str) and cell.value.strip():  # 检查单元格内容是否为空白字符串
                    translated_value = _translate_cell_value(cell, translation_core, source_lang, target_lang)
                    cell.value = translated_value
                    
                    # 更新单元格翻译进度
                    cell_progress = ((row_idx - 1) * original_sheet.max_column + cell_idx) / total_cells * 100
                    if task is not None:
                        task.update_state(
                            state='PROGRESS',
                            meta={
                                'current': sheet_idx,
                                'total': total_sheets,
                                'cell_progress': cell_progress,
                                'sheet_progress': (sheet_idx - 1) / total_sheets * 100,
                                'progress': ((sheet_idx - 1) + cell_progress / 100) / total_sheets * 100
                            }
                        )
        
        # 更新任务进度
        sheet_progress = (sheet_idx / total_sheets) * 100
        if task is not None:
            task.update_state(
                state='PROGRESS',
                meta={
                    'current': sheet_idx,
                    'total': total_sheets,
                    'cell_progress': 100,
                    'sheet_progress': sheet_progress,
                    'progress': sheet_progress
                }
            )
        logging.info(f"Progress: {sheet_progress:.1f}% - Translated sheet '{sheet_name}'")

    # 保存翻译后的工作簿
    excel_wb.save(output_path)
    logging.info(f"Completed translation. Saved to: {output_path}")

def _translate_cell_value(cell, translator, src_lang, tgt_lang):
    """翻译单元格文本内容"""
    try:
        return translator.translate_text(cell.value, src_lang, tgt_lang) if isinstance(cell.value, str) else cell.value
    except Exception as e:
        logging.error(f"Translation error at {cell.coordinate}: {str(e)}")
        return cell.value  # 出错时返回原值
